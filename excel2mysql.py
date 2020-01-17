import datetime
import os

from openpyxl import load_workbook, Workbook
import csv
import migration_dao as md


def migrate(filename, is_truncate=False, is_create_table=False, sheet_name='' ):
    fname, ext = os.path.splitext(filename)
    print(fname)
    print(ext)

    if ext == '.csv':
        global is_csv_file
        is_csv_file = True
        _temp_excel_filename = convert_csv_to_xlsx(filename, sheet_name)
        excel_filename = _temp_excel_filename
    elif ext == ".xlsx" or ext == ".xls":
        excel_filename = filename
    else:
        raise ValueError

    load_wb = load_workbook(excel_filename, data_only=True)
    dao = md.Dao()

    target_worshsheet = ''
    total_records_count = 0
    total_columns_count = 0
    total_columns_list = ''

    for _sheet_name in load_wb.sheetnames:

        target_worshsheet = _sheet_name
        load_ws = load_wb[_sheet_name]
        _records = []

        # truncate == True, 기존 테이블을 비우고, 다시 입력한다.
        if is_truncate:
            dao.trucate(_sheet_name)

        for _rows in load_ws.rows:
            _record = []
            for _cols in _rows:
                _record.append('' if _cols.value == None else _cols.value)

            _records.append(_record)

        _titles = _records[0]

        # title column 분리
        del _records[0]
        _values = _records

        total_records_count = len(_values)
        total_columns_count = len(_titles)
        total_columns_list = ", ".join(_titles)

        # sheet명으로 테이블 생성
        if is_create_table:
            now_dt = datetime.datetime.now()
            prefix = now_dt.strftime("%Y%m%d%H%M%S")
            if not dao.is_exist_table(_sheet_name):
                dao.create_table(_sheet_name, _titles)
            else:
                _sheet_name = prefix + "_" + _sheet_name
                dao.create_table(_sheet_name, _titles)
                # exit()

        _insert_keys = ", ".join(_titles)

        for idx, _v in enumerate(_values):

            _result = 0
            # _insert_values = "'" + "','".join(_v) +     "'"
            _values_str = ['%s' for i in _v]

            # query_sql = 'insert into ' + _sheet_name + '(' + _insert_keys + ')' + ' values (' + _insert_values + ')'
            query_sql = 'insert into ' + _sheet_name + '(' + _insert_keys + ')' + ' values (' + ", ".join(
                _values_str) + ')'
            _result = dao.insert(query_sql, _v)
            print("  " + str(idx) + " ==> " + query_sql)

            # database에 insert가 성공하면, 결과는 1, 실패는 0
            # 실패할경우 error를 표시
            if _result == 0:
                print('error index ==> ' + str(idx))

    result_report(target_worshsheet, total_columns_count, total_columns_list, total_records_count)


def convert_csv_to_xlsx(csv_filename, separator=",", sheet_name=''):
    fname, ext = os.path.splitext(csv_filename)

    wb = Workbook()
    sheet = wb.active
    if not sheet_name == '':
        sheet.title = sheet_name

    with open(csv_filename, 'r', encoding='UTF8') as f:
        for row in csv.reader(f):
            sheet.append(row)

    temp_excel_filename = "temp/" + "tmp_" + fname + ".xlsx"
    wb.save(temp_excel_filename)

    return temp_excel_filename


def result_report(target_worshsheet, total_columns_count, total_columns_list, total_records_count):
    print('')
    print(
        '|-- [RESULT_REPORT] -------------------------------------------------------------------------------------------------------------------------------------------------------------')
    print('|# target worksheet    => [' + target_worshsheet + ']')
    print('|# total columns list  => [' + total_columns_list + ']')
    print('|# total columns count => [' + str(total_columns_count) + ']')
    print('|# total records count => [' + str(total_records_count) + ']')
    print(
        '|--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')


if __name__ == '__main__':
    migrate('test.xlsx')
